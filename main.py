import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QLabel,
    QLineEdit, QPushButton, QFileDialog, QMessageBox,
)
from openpyxl import load_workbook
from PyQt5.QtGui import QIcon, QFont, QColor
from PyQt5.QtCore import Qt

class DocumentCheckerApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Document Register Checker")
        self.setGeometry(100, 100, 400, 350)
        self.setWindowIcon(QIcon('arup_icon.png'))  # Replace with Arup's icon file path

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        layout = QVBoxLayout()

        self.title_label = QLabel("Document Register Checker")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setFont(QFont("Arial", 24, QFont.Bold))
        layout.addWidget(self.title_label)

        self.excel_file_input = QLineEdit()
        self.excel_file_button = QPushButton("Browse Excel File")  # Added this button
        self.setup_button_style(self.excel_file_button)
        layout.addWidget(self.excel_file_input)
        layout.addWidget(self.excel_file_button)

        self.folder_location_input = QLineEdit()
        self.folder_location_button = QPushButton("Browse Folder Location")  # Added this button
        self.setup_button_style(self.folder_location_button)
        layout.addWidget(self.folder_location_input)
        layout.addWidget(self.folder_location_button)

        self.output_file_input = QLineEdit()
        self.output_file_button = QPushButton("Select Output Folder")  # Added this button
        self.setup_button_style(self.output_file_button)
        layout.addWidget(self.output_file_input)
        layout.addWidget(self.output_file_button)

        self.check_button = QPushButton("Check Documents")
        self.check_button.setStyleSheet("background-color: #E30613; color: white;")
        self.check_button.clicked.connect(self.check_documents)
        layout.addWidget(self.check_button)

        self.central_widget.setLayout(layout)

        self.excel_file_button.clicked.connect(self.browse_excel_file)
        self.folder_location_button.clicked.connect(self.browse_folder_location)
        self.output_file_button.clicked.connect(self.select_output_file_location)

    def setup_button_style(self, button):
        button.setStyleSheet("background-color: #E30613; color: white;")

    def setup_textfield_with_button(self, label_text, text_input, button, layout):
        text_label = QLabel(label_text)
        text_input.setStyleSheet("background-color: #F5F5F5;")
        button.setStyleSheet("background-color: #E30613; color: white;")

        text_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(text_label)
        layout.addWidget(text_input)
        layout.addWidget(button)

    def browse_excel_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        excel_file, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        self.excel_file_input.setText(excel_file)

    def browse_folder_location(self):
        folder_location = QFileDialog.getExistingDirectory(self, "Select Folder Location")
        self.folder_location_input.setText(folder_location)

    def select_output_file_location(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        output_file_dir = QFileDialog.getExistingDirectory(self, "Select Output File Location")
        self.output_file_input.setText(output_file_dir)

    def check_documents(self):
        excel_file_path = self.excel_file_input.text()
        folder_location = self.folder_location_input.text()
        output_file_dir = self.output_file_input.text()

        if excel_file_path and folder_location and output_file_dir:
            try:
                wb = load_workbook(excel_file_path)
                sheet = wb.active

                sheet.cell(row=1, column=sheet.max_column+1, value="Found in Folder")

                for row in sheet.iter_rows(min_row=2):
                    document_title = row[0].value
                    found = "N"
                    if self.check_document_in_folder(document_title, folder_location):
                        found = "Y"
                    sheet.cell(row=row[0].row, column=sheet.max_column, value=found)

                output_filename = os.path.join(output_file_dir, "document_checker_results.xlsx")
                wb.save(output_filename)
                wb.close()

                print("Document checking completed and results saved to:", output_filename)

                # Show a pop-up message when the check is complete
                QMessageBox.information(self, "Document Check Complete", "Document checking is complete. Results saved to:\n" + output_filename)

                # Open button to open the results file
                open_button = QPushButton("Open Results File")
                open_button.setStyleSheet("background-color: #2ECC71; color: white;")
                open_button.clicked.connect(lambda: os.startfile(output_filename))

                self.central_widget.layout().addWidget(open_button)

            except Exception as e:
                print("Error:", e)
        else:
            print("Please fill in all the required fields.")

    def check_document_in_folder(self, document_title, folder_location):
        for root, dirs, files in os.walk(folder_location):
            for filename in files:
                if document_title in filename:
                    return True
        return False

def main():
    app = QApplication(sys.argv)
    window = DocumentCheckerApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
